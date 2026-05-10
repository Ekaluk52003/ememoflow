from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django import forms
from django.http import HttpResponse
from django.contrib import messages
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from datetime import datetime
import json
import re
import openpyxl

from .models import Document, ApprovalWorkflow, DynamicField


def superuser_required(view_func):
    from functools import wraps
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if not request.user.is_superuser:
            return render(request, '403.html', status=403)
        return view_func(request, *args, **kwargs)
    return _wrapped_view


def _build_excel_response(queryset, workflow=None):
    """Build and return an Excel HttpResponse from a Document queryset."""
    queryset = queryset.select_related(
        'submitted_by', 'workflow', 'current_step', 'voided_by'
    ).prefetch_related('dynamic_values__field')

    if workflow:
        dynamic_fields = DynamicField.objects.filter(
            workflow=workflow
        ).order_by('order')
    else:
        dynamic_field_ids = set()
        field_id_to_name = {}
        for doc in queryset:
            for dv in doc.dynamic_values.all():
                dynamic_field_ids.add(dv.field.id)
                field_id_to_name[dv.field.id] = dv.field.name
        field_orders = {
            f.id: f.order for f in DynamicField.objects.filter(id__in=dynamic_field_ids).only('id', 'order')
        }
        dynamic_fields = [
            {'name': field_id_to_name[fid], 'order': field_orders.get(fid, 0)}
            for fid in field_id_to_name
        ]
        dynamic_fields.sort(key=lambda x: x['order'])

    doc_headers = [
        'Document Reference', 'Title', 'Content', 'Submitted By', 'Workflow',
        'Current Step', 'Status', 'Created At', 'Updated At',
        'Last Submitted At', 'Void Reason', 'Voided At', 'Voided By',
    ]

    if workflow:
        sorted_dynamic_field_names = [f.name for f in dynamic_fields]
    else:
        sorted_dynamic_field_names = [f['name'] for f in dynamic_fields]

    headers = doc_headers + sorted_dynamic_field_names

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Documents'

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    def format_value(val):
        if val is None:
            return ''
        if isinstance(val, datetime):
            return val.replace(tzinfo=None)
        return val

    def strip_html(html):
        if not html:
            return ''
        text = re.sub(r'<[^>]+>', ' ', html)
        text = re.sub(r'&nbsp;', ' ', text)
        text = re.sub(r'\s+', ' ', text).strip()
        return text

    json_columns = set()

    def format_product_list(products):
        lines = []
        for idx, product in enumerate(products, 1):
            parts = []
            if product.get('code'):
                parts.append(f"Code: {product['code']}")
            if product.get('name'):
                parts.append(f"Name: {product['name']}")
            if product.get('quantity') is not None:
                parts.append(f"Qty: {product['quantity']}")
            if product.get('unit'):
                parts.append(f"Unit: {product['unit']}")
            if product.get('batch_no'):
                parts.append(f"Batch: {product['batch_no']}")
            if product.get('bin_location'):
                parts.append(f"Bin: {product['bin_location']}")
            lines.append(f"{idx}. {' | '.join(parts)}")
        return '\n'.join(lines)

    def format_table_list(field, rows):
        columns = []
        table_columns = getattr(field, 'table_columns', '') or ''
        for spec in table_columns.split('|'):
            spec = spec.strip()
            if not spec:
                continue
            label = spec.split(':', 1)[0].strip()
            if label:
                columns.append(label)

        if not columns or not rows:
            return ''

        lines = []
        header_line = '  |  '.join(columns)
        lines.append(header_line)
        lines.append('-' * len(header_line))
        for row in rows:
            row_values = []
            for col in columns:
                val = row.get(col, '')
                if val is None:
                    val = ''
                row_values.append(str(val))
            lines.append('  |  '.join(row_values))
        return '\n'.join(lines)

    for row_num, doc in enumerate(queryset, 2):
        dyn_values = {}
        for dv in doc.dynamic_values.all():
            if dv.field.field_type == 'attachment' and dv.file:
                dyn_values[dv.field.name] = dv.file.name
            elif dv.field.field_type == 'product_list' and dv.json_value:
                products = dv.json_value if isinstance(dv.json_value, list) else []
                dyn_values[dv.field.name] = format_product_list(products)
                json_columns.add(dv.field.name)
            elif dv.field.field_type == 'table_list' and dv.json_value:
                rows = dv.json_value if isinstance(dv.json_value, list) else []
                dyn_values[dv.field.name] = format_table_list(dv.field, rows)
                json_columns.add(dv.field.name)
            else:
                val = dv.value or ''
                if dv.field.field_type in ('tiptap_editor', 'textarea', 'text'):
                    val = strip_html(val)
                dyn_values[dv.field.name] = val

        row_data = [
            doc.document_reference,
            doc.title,
            strip_html(doc.content),
            doc.submitted_by.get_full_name() or doc.submitted_by.username if doc.submitted_by else '',
            doc.workflow.name if doc.workflow else '',
            doc.current_step.name if doc.current_step else '',
            doc.get_status_display() if hasattr(doc, 'get_status_display') else doc.status,
            format_value(doc.created_at),
            format_value(doc.updated_at),
            format_value(doc.last_submitted_at),
            doc.void_reason or '',
            format_value(doc.voided_at),
            doc.voided_by.get_full_name() or doc.voided_by.username if doc.voided_by else '',
        ] + [dyn_values.get(name, '') for name in sorted_dynamic_field_names]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            if headers[col_num - 1] in json_columns:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        max_length = len(str(header))
        for row in ws.iter_rows(min_row=2, min_col=col_num, max_col=col_num):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 60)
        if headers[col_num - 1] in json_columns:
            adjusted_width = max(adjusted_width, 40)
        ws.column_dimensions[col_letter].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"documents_export_{workflow.name.replace(' ', '_')}.xlsx" if workflow else "documents_export.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


class DocumentExportForm(forms.Form):
    workflow = forms.ModelChoiceField(
        queryset=ApprovalWorkflow.objects.all().order_by('name'),
        label='Workflow',
        required=True,
        empty_label='-- Choose a workflow --',
    )
    status = forms.ChoiceField(
        label='Status',
        required=False,
        choices=[('', '-- All Statuses --')] + list(Document.STATUS_CHOICES),
    )
    date_from = forms.DateField(
        label='Created From',
        required=False,
        widget=forms.DateInput(attrs={'type': 'date'}),
    )
    date_to = forms.DateField(
        label='Created To',
        required=False,
        widget=forms.DateInput(attrs={'type': 'date'}),
    )


MAX_EXPORT_ROWS = 1000


@login_required
@superuser_required
def export_documents(request):
    """Superuser-only page to export documents to Excel with workflow filters."""
    count = None
    can_export = False

    if request.method == 'POST':
        form = DocumentExportForm(request.POST)
        if form.is_valid():
            workflow = form.cleaned_data['workflow']
            status = form.cleaned_data['status']
            date_from = form.cleaned_data['date_from']
            date_to = form.cleaned_data['date_to']

            qs = Document.objects.filter(workflow=workflow)
            if status:
                qs = qs.filter(status=status)
            if date_from:
                qs = qs.filter(created_at__date__gte=date_from)
            if date_to:
                qs = qs.filter(created_at__date__lte=date_to)

            count = qs.count()

            action_type = request.POST.get('action_type', '')
            is_htmx = request.headers.get('HX-Request') == 'true'

            if action_type == 'export':
                if count == 0:
                    messages.error(request, 'No documents found for the selected filters.')
                elif count > MAX_EXPORT_ROWS:
                    messages.error(
                        request,
                        f'This export would include {count} documents. The maximum is {MAX_EXPORT_ROWS}. '
                        'Please add more filters (e.g. a narrower date range or status) to narrow the results.'
                    )
                else:
                    return _build_excel_response(qs, workflow=workflow)
            else:
                # Preview mode
                can_export = 0 < count <= MAX_EXPORT_ROWS
                if is_htmx:
                    return render(request, 'document/partials/export_count_result.html', {
                        'count': count,
                        'can_export': can_export,
                        'max_rows': MAX_EXPORT_ROWS,
                    })
    else:
        form = DocumentExportForm()

    return render(request, 'document/export_documents.html', {
        'form': form,
        'count': count,
        'can_export': can_export,
        'max_rows': MAX_EXPORT_ROWS,
    })
